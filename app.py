import streamlit as st
import pandas as pd
from openpyxl import load_workbook
import datetime
import joblib
import numpy as np
import nltk
import dossier_utils as utils

# --- Configuración de la página ---
st.set_page_config(
    page_title="Dossier Intelligence · Nissan",
    page_icon="📰",
    layout="wide",
    initial_sidebar_state="collapsed"
)

# --- Estilos CSS modernos: blanco con acentos teal/verde profesional ---
st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=DM+Sans:wght@300;400;500;600;700&family=DM+Mono:wght@400;500&display=swap');

    html, body, [class*="css"] {
        font-family: 'DM Sans', sans-serif;
    }

    .stApp {
        background-color: #F7F9F8;
    }

    #MainMenu, footer, header { visibility: hidden; }
    .block-container {
        padding-top: 2rem;
        padding-bottom: 3rem;
        max-width: 1100px;
    }

    /* ── HEADER ── */
    .app-header {
        background: linear-gradient(135deg, #0D4F3C 0%, #1A7A5E 60%, #22A37A 100%);
        border-radius: 16px;
        padding: 2.5rem 2.8rem;
        margin-bottom: 2rem;
        position: relative;
        overflow: hidden;
    }
    .app-header::before {
        content: '';
        position: absolute;
        top: -40px; right: -40px;
        width: 200px; height: 200px;
        background: rgba(255,255,255,0.06);
        border-radius: 50%;
    }
    .app-header::after {
        content: '';
        position: absolute;
        bottom: -60px; left: 30%;
        width: 300px; height: 180px;
        background: rgba(255,255,255,0.04);
        border-radius: 50%;
    }
    .app-header h1 {
        color: #FFFFFF;
        font-size: 1.9rem;
        font-weight: 700;
        letter-spacing: -0.02em;
        margin: 0 0 0.3rem 0;
    }
    .app-header p {
        color: rgba(255,255,255,0.75);
        font-size: 0.95rem;
        font-weight: 400;
        margin: 0;
    }
    .app-header .badge {
        display: inline-block;
        background: rgba(255,255,255,0.15);
        color: #fff;
        font-size: 0.72rem;
        font-weight: 600;
        letter-spacing: 0.08em;
        text-transform: uppercase;
        padding: 0.25rem 0.7rem;
        border-radius: 20px;
        border: 1px solid rgba(255,255,255,0.25);
        margin-bottom: 0.8rem;
    }

    /* ── TARJETAS ── */
    .card {
        background: #FFFFFF;
        border: 1px solid #E8EFEC;
        border-radius: 12px;
        padding: 1.6rem 1.8rem;
        margin-bottom: 1.2rem;
        box-shadow: 0 1px 4px rgba(13,79,60,0.06);
    }
    .card-title {
        font-size: 0.8rem;
        font-weight: 600;
        letter-spacing: 0.1em;
        text-transform: uppercase;
        color: #1A7A5E;
        margin-bottom: 1rem;
    }

    /* ── MÉTRICAS ── */
    .metrics-row {
        display: flex;
        gap: 1rem;
        margin: 1.5rem 0;
    }
    .metric-card {
        flex: 1;
        background: #FFFFFF;
        border: 1px solid #E8EFEC;
        border-radius: 12px;
        padding: 1.4rem 1.6rem;
        text-align: center;
        box-shadow: 0 1px 4px rgba(13,79,60,0.06);
    }
    .metric-card .metric-value {
        font-size: 2.2rem;
        font-weight: 700;
        color: #0D4F3C;
        line-height: 1;
        margin-bottom: 0.4rem;
        font-family: 'DM Mono', monospace;
    }
    .metric-card .metric-label {
        font-size: 0.78rem;
        color: #6B8F82;
        font-weight: 500;
        letter-spacing: 0.03em;
    }
    .metric-card.accent .metric-value { color: #1A7A5E; }
    .metric-card.muted .metric-value { color: #9BB5AC; }

    /* ── BANNER DE ÉXITO ── */
    .success-banner {
        background: linear-gradient(135deg, #0D4F3C, #1A7A5E);
        border-radius: 12px;
        padding: 1.2rem 1.8rem;
        display: flex;
        align-items: center;
        gap: 1rem;
        margin: 1.5rem 0;
        box-shadow: 0 2px 12px rgba(13,79,60,0.2);
    }
    .success-banner .icon {
        font-size: 1.6rem;
        line-height: 1;
    }
    .success-banner .text strong {
        display: block;
        color: #FFFFFF;
        font-size: 1rem;
        font-weight: 700;
        margin-bottom: 0.15rem;
    }
    .success-banner .text span {
        color: rgba(255,255,255,0.72);
        font-size: 0.85rem;
    }

    /* ── STEPS ── */
    .step {
        display: flex;
        align-items: flex-start;
        gap: 1rem;
        margin-bottom: 0.8rem;
    }
    .step-num {
        min-width: 28px;
        height: 28px;
        background: #1A7A5E;
        color: white;
        border-radius: 50%;
        display: flex;
        align-items: center;
        justify-content: center;
        font-size: 0.78rem;
        font-weight: 700;
        margin-top: 1px;
    }
    .step-text { color: #2D5A4A; font-size: 0.9rem; line-height: 1.6; }

    /* ── FILE UPLOADER ── */
    [data-testid="stFileUploader"] {
        background: #FFFFFF;
        border: 2px dashed #B2D4C8;
        border-radius: 12px;
        padding: 0.5rem;
        transition: border-color 0.2s;
    }
    [data-testid="stFileUploader"]:hover {
        border-color: #1A7A5E;
    }

    /* ── BOTÓN PRIMARIO ── */
    .stButton > button[kind="primary"] {
        background: linear-gradient(135deg, #0D4F3C, #1A7A5E);
        color: white;
        border: none;
        border-radius: 10px;
        padding: 0.7rem 2rem;
        font-family: 'DM Sans', sans-serif;
        font-size: 0.95rem;
        font-weight: 600;
        letter-spacing: 0.01em;
        transition: all 0.2s;
        box-shadow: 0 2px 8px rgba(13,79,60,0.25);
        width: 100%;
        height: 48px;
    }
    .stButton > button[kind="primary"]:hover {
        background: linear-gradient(135deg, #0A3D2E, #156A50);
        box-shadow: 0 4px 14px rgba(13,79,60,0.35);
        transform: translateY(-1px);
    }
    .stButton > button[kind="primary"]:disabled {
        background: #C5D9D4;
        box-shadow: none;
        transform: none;
    }

    /* ── BOTÓN DESCARGA ── */
    .stDownloadButton > button {
        background: #FFFFFF;
        color: #0D4F3C;
        border: 2px solid #1A7A5E;
        border-radius: 10px;
        padding: 0.65rem 1.8rem;
        font-family: 'DM Sans', sans-serif;
        font-size: 0.9rem;
        font-weight: 600;
        transition: all 0.2s;
        width: 100%;
        height: 48px;
    }
    .stDownloadButton > button:hover {
        background: #F0F7F4;
        box-shadow: 0 2px 10px rgba(13,79,60,0.15);
        transform: translateY(-1px);
    }

    /* ── ALERTAS ── */
    .stSuccess {
        background: #F0F7F4;
        border: 1px solid #B2D4C8;
        border-radius: 10px;
        color: #0D4F3C;
    }
    .stWarning {
        background: #FFF8ED;
        border: 1px solid #F5C97A;
        border-radius: 10px;
    }
    .stError {
        background: #FEF2F2;
        border: 1px solid #FCA5A5;
        border-radius: 10px;
    }

    /* ── PROGRESS BAR ── */
    .stProgress > div > div > div {
        background: linear-gradient(90deg, #1A7A5E, #22A37A);
        border-radius: 4px;
    }

    /* ── EXPANDER ── */
    .streamlit-expanderHeader {
        background: #F7F9F8;
        border-radius: 8px;
        font-weight: 500;
        color: #0D4F3C;
    }

    /* ── DATAFRAME ── */
    [data-testid="stDataFrame"] {
        border-radius: 10px;
        overflow: hidden;
        border: 1px solid #E8EFEC;
    }

    hr {
        border: none;
        border-top: 1px solid #E8EFEC;
        margin: 1.5rem 0;
    }

    .stSpinner > div { border-top-color: #1A7A5E !important; }

    /* ── FILE STATUS ── */
    .file-status {
        display: flex;
        align-items: center;
        gap: 0.6rem;
        padding: 0.7rem 1rem;
        border-radius: 8px;
        font-size: 0.87rem;
        font-weight: 500;
        margin-top: 0.5rem;
    }
    .file-status.ok {
        background: #F0F7F4;
        color: #0D4F3C;
        border: 1px solid #B2D4C8;
    }
    .file-status.missing {
        background: #FFF8ED;
        color: #92400E;
        border: 1px solid #F5C97A;
    }

    .results-header {
        font-size: 1.1rem;
        font-weight: 700;
        color: #0D4F3C;
        margin: 1.8rem 0 1rem 0;
        padding-bottom: 0.5rem;
        border-bottom: 2px solid #E8EFEC;
    }
</style>
""", unsafe_allow_html=True)

# --- Descarga NLTK stopwords si es necesario ---
try:
    nltk.data.find('corpora/stopwords')
except LookupError:
    with st.spinner("Descargando recursos de lenguaje por primera vez..."):
        nltk.download('stopwords')


# ==============================================================================
# FUNCIONES DE CARGA Y PROCESO PRINCIPAL
# ==============================================================================

@st.cache_resource
def load_ml_models():
    """Carga los pipelines completos de sentimiento y tema."""
    try:
        sentiment_pipeline = joblib.load('pipeline_sentimiento.pkl')
        topic_pipeline = joblib.load('pipeline_tema.pkl')
        return sentiment_pipeline, topic_pipeline
    except FileNotFoundError as e:
        st.error(
            f"**Error Crítico:** No se encontró `{e.filename}`. "
            "Asegúrate de que `pipeline_sentimiento.pkl` y `pipeline_tema.pkl` "
            "estén en la misma carpeta que esta app."
        )
        st.stop()


def read_and_expand_dossier(dossier_file):
    """Lee el archivo Excel, extrae hyperlinks y expande las filas por mención."""
    wb = load_workbook(dossier_file)
    sheet = wb.active
    original_headers = [cell.value for cell in sheet[1] if cell.value]

    rows_data = []
    for row in sheet.iter_rows(min_row=2, values_only=False):
        if all(c.value is None for c in row):
            continue

        row_values = {}
        for i, header in enumerate(original_headers):
            if header in ['Link Nota', 'Link (Streaming - Imagen)']:
                row_values[header] = utils.extract_link_from_cell(row[i])
            else:
                row_values[header] = row[i].value

        menciones_str = str(row_values.get('Menciones - Empresa') or '')
        menciones = [m.strip() for m in menciones_str.split(';') if m.strip()]

        if not menciones:
            rows_data.append(row_values)
        else:
            for mencion in menciones:
                new_row = row_values.copy()
                new_row['Menciones - Empresa'] = mencion
                rows_data.append(new_row)

    return pd.DataFrame(rows_data)


def run_full_process(dossier_file, config_file):
    st.markdown("<hr>", unsafe_allow_html=True)
    progress_bar = st.progress(0, text="Iniciando proceso...")

    # --- 1. Carga de modelos y configuración ---
    progress_bar.progress(5, text="Paso 1 / 8 — Cargando modelos y configuración...")
    sentiment_pipeline, topic_pipeline = load_ml_models()

    try:
        config_sheets = pd.read_excel(config_file, sheet_name=None)
        region_map = pd.Series(
            config_sheets['Regiones'].iloc[:, 1].values,
            index=config_sheets['Regiones'].iloc[:, 0].astype(str).str.lower().str.strip()
        ).to_dict()
        internet_map = pd.Series(
            config_sheets['Internet'].iloc[:, 1].values,
            index=config_sheets['Internet'].iloc[:, 0].astype(str).str.lower().str.strip()
        ).to_dict()
        mention_map = pd.Series(
            config_sheets['Menciones'].iloc[:, 1].values,
            index=config_sheets['Menciones'].iloc[:, 0].astype(str).str.strip()
        ).to_dict()
        final_topic_map = pd.Series(
            config_sheets['Mapa_Temas'].iloc[:, 1].values,
            index=config_sheets['Mapa_Temas'].iloc[:, 0].astype(str).str.strip()
        ).to_dict()
    except Exception as e:
        st.error(f"**Error al cargar `Configuracion.xlsx`:** {e}")
        st.stop()

    # --- 2. Lectura y Expansión del Dossier ---
    progress_bar.progress(15, text="Paso 2 / 8 — Leyendo Dossier y expandiendo filas...")
    df = read_and_expand_dossier(dossier_file)

    # --- 3. Limpieza y Normalización de Datos ---
    progress_bar.progress(25, text="Paso 3 / 8 — Aplicando mapeos y normalizaciones...")
    if 'Fecha' in df.columns:
        df['Fecha'] = pd.to_datetime(df['Fecha'], errors='coerce', dayfirst=True)
        if df['Fecha'].isna().any():
            st.warning("⚠️ Algunas fechas no se pudieron convertir. Revisa el archivo original.")

    if 'Título' in df.columns:
        df['Título'] = df['Título'].apply(utils.clean_title)

    if 'Resumen - Aclaracion' in df.columns:
        df['Resumen - Aclaracion'] = df['Resumen - Aclaracion'].apply(utils.corregir_resumen)

    tipo_medio_map = {
        'online': 'Internet', 'diario': 'Prensa',
        'am': 'Radio', 'fm': 'Radio',
        'aire': 'Televisión', 'cable': 'Televisión',
        'revista': 'Revista'
    }
    df['Tipo de Medio'] = (
        df['Tipo de Medio'].str.lower().str.strip()
        .map(tipo_medio_map)
        .fillna(df['Tipo de Medio'])
    )

    if 'Medio' in df.columns:
        df['Región'] = df['Medio'].astype(str).str.lower().str.strip().map(region_map)

    if 'Menciones - Empresa' in df.columns:
        df['Menciones - Empresa'] = (
            df['Menciones - Empresa'].astype(str).str.strip()
            .map(mention_map)
            .fillna(df['Menciones - Empresa'])
        )

    is_internet = df['Tipo de Medio'] == 'Internet'
    if is_internet.any():
        df.loc[is_internet, 'Medio'] = (
            df.loc[is_internet, 'Medio'].astype(str).str.lower().str.strip()
            .map(internet_map)
            .fillna(df.loc[is_internet, 'Medio'])
        )

    # --- 4. Reorganización de Columnas ---
    progress_bar.progress(40, text="Paso 4 / 8 — Reorganizando columnas de links y dimensiones...")
    is_print = df['Tipo de Medio'].isin(['Prensa', 'Revista'])
    is_broadcast = df['Tipo de Medio'].isin(['Radio', 'Televisión'])

    if 'Link Nota' in df.columns and 'Link (Streaming - Imagen)' in df.columns:
        df.loc[is_internet, ['Link Nota', 'Link (Streaming - Imagen)']] = (
            df.loc[is_internet, ['Link (Streaming - Imagen)', 'Link Nota']].values
        )
        cond_copy = is_print & df['Link Nota'].isnull() & df['Link (Streaming - Imagen)'].notnull()
        df.loc[cond_copy, 'Link Nota'] = df.loc[cond_copy, 'Link (Streaming - Imagen)']
        df.loc[is_print | is_broadcast, 'Link (Streaming - Imagen)'] = None

    if 'Duración - Nro. Caracteres' in df.columns and 'Dimensión' in df.columns:
        df.loc[is_broadcast, 'Dimensión'] = df.loc[is_broadcast, 'Duración - Nro. Caracteres']
        df.loc[is_broadcast, 'Duración - Nro. Caracteres'] = np.nan

    # --- 5. Detección de Duplicados ---
    progress_bar.progress(50, text="Paso 5 / 8 — Detectando duplicados...")
    df = utils.detect_duplicates_optimized(df)

    # --- 6. Aplicación de Modelos de IA ---
    progress_bar.progress(70, text="Paso 6 / 8 — Aplicando modelos de IA a noticias únicas...")
    df_valid = df[~df['is_duplicate']].copy()
    if not df_valid.empty:
        df_valid['texto_para_ia'] = (
            df_valid['Título'].fillna('') + ' ' + df_valid['Resumen - Aclaracion'].fillna('')
        )

        preds_sent = sentiment_pipeline.predict(df_valid['texto_para_ia'])
        label_map_inv = {1: 'Positivo', 0: 'Neutro', -1: 'Negativo'}
        df_valid['Tono'] = [label_map_inv.get(p, 'Indefinido') for p in preds_sent]

        df_valid['resumen_procesado'] = df_valid['texto_para_ia'].apply(
            utils.preprocess_text_for_topic
        )
        df_valid['Temas Generales - Tema'] = topic_pipeline.predict(df_valid['resumen_procesado'])

        df.update(df_valid[['Tono', 'Temas Generales - Tema']])

    # --- 7. Homogeneización de Temas y Mapeo Final ---
    progress_bar.progress(85, text="Paso 7 / 8 — Homogeneizando y mapeando temas...")
    df_valid_homog = df[~df['is_duplicate']].copy()
    if not df_valid_homog.empty and 'Temas Generales - Tema' in df_valid_homog.columns:
        df_valid_homog['titulo_norm_homog'] = df_valid_homog['Título'].apply(
            utils.normalize_title_for_comparison
        )
        homogenized_temas = df_valid_homog.groupby(
            'titulo_norm_homog'
        )['Temas Generales - Tema'].transform(
            lambda x: x.mode()[0] if not x.mode().empty else x
        )
        df_valid_homog['Temas Generales - Tema'] = homogenized_temas
        df.update(df_valid_homog[['Temas Generales - Tema']])

    if 'Temas Generales - Tema' in df.columns:
        df['Tema'] = (
            df['Temas Generales - Tema'].astype(str).str.strip()
            .map(final_topic_map)
            .fillna('Indefinido')
        )

    # --- Marcado final de duplicadas ---
    mask_dup = df['is_duplicate']
    if mask_dup.any():
        if 'Temas Generales - Tema' in df.columns:
            df.loc[mask_dup, 'Temas Generales - Tema'] = '-'
        if 'Tema' in df.columns:
            df.loc[mask_dup, 'Tema'] = '-'
        df.loc[mask_dup, 'Tono'] = 'Duplicada'

    # --- 8. Resultados Finales ---
    progress_bar.progress(100, text="✓ Proceso completado")

    final_order = [
        "ID Noticia", "Fecha", "Hora", "Medio", "Tipo de Medio", "Sección - Programa",
        "Región", "Título", "Autor - Conductor", "Nro. Pagina", "Dimensión",
        "Duración - Nro. Caracteres", "CPE", "Tier", "Audiencia", "Tono", "Tema",
        "Temas Generales - Tema", "Resumen - Aclaracion", "Link Nota",
        "Link (Streaming - Imagen)", "Menciones - Empresa"
    ]

    total = len(df)
    dups_count = int(mask_dup.sum())
    unique_count = total - dups_count
    filename = f"Dossier_Procesado_{datetime.datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    excel_data = utils.to_excel_from_df(df, final_order)

    # Banner de éxito
    st.markdown(f"""
    <div class="success-banner">
        <div class="icon">⚡</div>
        <div class="text">
            <strong>Proceso finalizado correctamente</strong>
            <span>{total:,} filas procesadas · {unique_count:,} únicas · {dups_count:,} duplicadas</span>
        </div>
    </div>
    """, unsafe_allow_html=True)

    # Métricas
    st.markdown('<p class="results-header">Resumen del proceso</p>', unsafe_allow_html=True)
    st.markdown(f"""
    <div class="metrics-row">
        <div class="metric-card">
            <div class="metric-value">{total:,}</div>
            <div class="metric-label">Filas totales procesadas</div>
        </div>
        <div class="metric-card accent">
            <div class="metric-value">{unique_count:,}</div>
            <div class="metric-label">Noticias únicas analizadas</div>
        </div>
        <div class="metric-card muted">
            <div class="metric-value">{dups_count:,}</div>
            <div class="metric-label">Filas marcadas como duplicadas</div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    # Botones: Iniciar (deshabilitado visualmente) | Descargar — lado a lado
    st.markdown('<p class="results-header">Descarga</p>', unsafe_allow_html=True)
    btn_col1, btn_col2 = st.columns(2)
    with btn_col1:
        st.button("✓ Proceso completado", disabled=True, type="primary")
    with btn_col2:
        st.download_button(
            label="⬇ Descargar archivo procesado (.xlsx)",
            data=excel_data,
            file_name=filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    # Previsualización
    st.markdown('<p class="results-header">Previsualización de resultados</p>', unsafe_allow_html=True)
    final_cols_in_df = [col for col in final_order if col in df.columns]
    df_display = df[final_cols_in_df].copy()

    if 'Fecha' in df_display.columns:
        df_display['Fecha'] = (
            pd.to_datetime(df_display['Fecha'])
            .dt.strftime('%d/%m/%Y')
            .replace('NaT', 'FECHA INVÁLIDA')
        )

    st.dataframe(df_display, use_container_width=True, hide_index=True)


# ==============================================================================
# INTERFAZ PRINCIPAL
# ==============================================================================

# Header
st.markdown("""
<div class="app-header">
    <div class="badge">Nissan · Media Intelligence</div>
    <h1>Procesador de Dossiers</h1>
    <p>Limpieza, enriquecimiento y análisis automático de dossieres de prensa · v4.1</p>
</div>
""", unsafe_allow_html=True)

# Instrucciones
st.markdown("""
<div class="card">
    <div class="card-title">Cómo usar esta herramienta</div>
    <div class="step">
        <div class="step-num">1</div>
        <div class="step-text">Prepara tu archivo <strong>Dossier</strong> (.xlsx) y el archivo <strong>Configuracion.xlsx</strong>.</div>
    </div>
    <div class="step">
        <div class="step-num">2</div>
        <div class="step-text">Sube ambos archivos en el área de carga de abajo. El sistema los detecta automáticamente.</div>
    </div>
    <div class="step">
        <div class="step-num">3</div>
        <div class="step-text">Haz clic en <strong>Iniciar proceso</strong> y espera a que finalice el análisis.</div>
    </div>
    <div class="step">
        <div class="step-num">4</div>
        <div class="step-text">Descarga el archivo final desde el botón que aparece al lado del estado del proceso.</div>
    </div>
</div>
""", unsafe_allow_html=True)

# Estructura de configuración
with st.expander("📋  Ver estructura requerida para Configuracion.xlsx"):
    st.markdown("""
    | Hoja | Columna A | Columna B |
    |------|-----------|-----------|
    | `Regiones` | Medio | Región |
    | `Internet` | Medio Original | Medio Mapeado |
    | `Menciones` | Mención Original | Mención Mapeada |
    | `Mapa_Temas` | Temas Generales - Tema | Tema |
    """)

st.markdown("<br>", unsafe_allow_html=True)

# Carga de archivos
st.markdown('<div class="card-title">Carga de archivos</div>', unsafe_allow_html=True)
uploaded_files = st.file_uploader(
    "Arrastra los archivos aquí o haz clic para seleccionarlos",
    type=["xlsx"],
    accept_multiple_files=True,
    label_visibility="collapsed"
)

dossier_file, config_file = None, None

if uploaded_files:
    for file in uploaded_files:
        if 'config' in file.name.lower():
            config_file = file
        else:
            dossier_file = file

    col_a, col_b = st.columns(2)
    with col_a:
        if dossier_file:
            st.markdown(
                f'<div class="file-status ok">✓ Dossier cargado — <strong>{dossier_file.name}</strong></div>',
                unsafe_allow_html=True
            )
        else:
            st.markdown(
                '<div class="file-status missing">⚠ No se detectó el archivo Dossier</div>',
                unsafe_allow_html=True
            )
    with col_b:
        if config_file:
            st.markdown(
                f'<div class="file-status ok">✓ Configuración cargada — <strong>{config_file.name}</strong></div>',
                unsafe_allow_html=True
            )
        else:
            st.markdown(
                '<div class="file-status missing">⚠ No se detectó Configuracion.xlsx</div>',
                unsafe_allow_html=True
            )

st.markdown("<br>", unsafe_allow_html=True)

# Botones principales: Iniciar | Descargar (descargar aparece solo tras el proceso)
col_start, col_download = st.columns(2)

with col_start:
    start_clicked = st.button(
        "▶  Iniciar proceso completo",
        disabled=not (dossier_file and config_file),
        type="primary"
    )

# El placeholder en col_download se rellena dentro de run_full_process
# tras generar el Excel. Para mantener el layout simétrico mostramos
# un botón inactivo mientras no se ha procesado nada.
with col_download:
    if not start_clicked:
        st.button(
            "⬇ Descargar archivo procesado (.xlsx)",
            disabled=True,
            type="primary"
        )

if start_clicked:
    run_full_process(dossier_file, config_file)
